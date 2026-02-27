import pandas as pd
import os

# Ruta del archivo
file_path = "data/MENU_MASTER_FINAL.xlsx"

if not os.path.exists(file_path):
    print("Archivo no encontrado.")
    exit()

print(f"Abriendo {file_path}...")

# Cargar Excel existente
# Usamos openpyxl engine para append mode o pandas replace
# Para simplificar y mantener formato, reescribimos los datos respetando columnas
# Primero leemos para no perder nada si hubiera (aunque está vacio salvo dummy)

# --- DATOS EXTRAÍDOS DE IMAGENES (OCR MANUAL) ---
data_comida = [
    # Categoria, Subcat, NombreES, NombreEN, DescES, DescEN, Precio, Alergenos, Visible
    # Pagina 1 / Imagen 1 (Entrantes implicitos / Ensaladas)
    ["ENTRANTES", "", "FIGATELL DE SEPIA CON PICADA DE CACAHUETE", "CUTTLEFISH MEATBALL WITH PEANUT DRESSING", "MAYONESA DE CHILES", "CHILLI MAYONNAISE", 13, "3, 4, 5, 6, 7, 10, 11, 12, 13", "SI"],
    ["ENTRANTES", "", "TITAINA DEL CABANYAL", "FRIED VEGETABLES WITH TOMATO", "HUEVO FRITO Y CHIPS DE PATATA", "FRIED EGG & POTATO CHIPS", 14, "3, 5, 6", "SI"],
    ["ENTRANTES", "", "TABLA DE QUESO", "CHEESE PLATE", "TOSTAS DE PAN Y FRUTOS SECOS", "BREAD TOASTS & NUTS", 18, "3, 4, 11", "SI"],
    ["ENTRANTES", "", "JAMÓN IBÉRICO DE BELLOTA CASTRO Y GONZÁLEZ (100GR)", "IBERIAN HAM", "", "", 25, "4", "SI"],
    
    # Ensaladas
    ["PRINCIPALES", "ENSALADAS", "ENSALADA DE TOMATE VALENCIANO", "VALENCIAN TOMATO SALAD", "VENTRESCA DE ATÚN Y CEBOLLA TIERNA", "TUNA BELLY & SPRING ONION", 13, "6", "SI"],
    ["PRINCIPALES", "ENSALADAS", "ENSALADA MEDITERRÁNEA", "MEDITERRANEAN SALAD", "TOMATE, HUEVO DURO Y ATÚN", "TOMATO, BOILED EGG & TUNA", 12, "5, 6", "SI"],
    ["PRINCIPALES", "ENSALADAS", "ENSALADA CÉSAR", "CESAR SALAD", "POLLO Y QUESO PARMESANO", "CHICKEN & PARMESAN CHEESE", 14, "4, 5, 6, 8, 11", "SI"],
    ["PRINCIPALES", "", "CREMA DEL CHEF", "CHEF'S CREAM", "VERDURAS DE TEMPORADA", "SEASONAL VEGETABLES", 11, "", "SI"],

    # Pagina 3 / Imagen 3 (Para Compartir)
    ["PARA COMPARTIR", "", "NACHOS", "NACHOS", "GUACAMOLE Y CREMA DE QUESO CHEDDAR", "GUACAMOLE & CHEDDAR CREAM CHEESE", 9, "2, 4, 11, 12", "SI"],
    ["PARA COMPARTIR", "", "BOQUERONES EN VINAGRE", "ANCHOVIES IN VINEGAR", "PAPAS, PEPINILLOS Y CEBOLLA ENCURTIDA", "CRISPS, PICKLES & PICKLED ONION", 10, "4, 6, 12", "SI"],
    ["PARA COMPARTIR", "", "CROQUETAS DE JAMÓN IBÉRICO", "IBERIAN HAM CROQUETTES", "", "", 11, "4, 5, 11", "SI"],
    ["PARA COMPARTIR", "", "CROQUETAS DE CARABINERO", "KING PRAWN CROQUETTES", "TOGARASHI Y MAYONESA JAPONESA", "TOGARASHI & JAPANESE MAYONNAISE", 13, "1, 4, 5, 6, 7, 8, 11, 13", "SI"],
    ["PARA COMPARTIR", "", "BUÑUELOS DE BACALAO EN TEMPURA", "TEMPURA COD FRITTER", "MERMELADA DE TOMATE Y ALIOLI", "TOMATO JAM & GARLIC MAYONNAISE", 12, "4, 5, 6", "SI"],
    ["PARA COMPARTIR", "", "COCA DE ACEITE CON SARDINA AHUMADA", "OIL CAKE WITH SMOKED SARDINE", "QUESO CURADO DE OVEJA", "CURED SHEEP'S CHEESE", 10, "4, 6, 11", "SI"],
    ["PARA COMPARTIR", "", "TORREZNOS DE SORIA", "FRIED PORK BELLY", "MOUSSE DE AGUACATE Y CEBOLLA ENCURTIDA", "AVOCADO MOUSSE & PICKLED ONION", 12, "11, 12", "SI"],
]

data_bebida = [
    # Categoria, Subcat, NombreES, NombreEN, DescES, DescEN, PrecioCopa, PrecioBot, Alergenos, Visible
    # Vinos Blancos (Imagen 2)
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "ENRIQUE MENDOZA, CHARDONNAY", "", "", "", "", 23, "", "SI"],
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "FINCA COLLADO, CHARDONNAY, MOSCATEL", "", "", "", "", 22, "", "SI"],
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "NODUS, SAUVIGNON BLANC", "", "", "", 4.5, 20, "", "SI"],
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "BOBAL BLANCO, BOBAL", "", "", "", "", 25, "", "SI"],
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "INVERSO, GODELLO", "", "", "", "", 26, "", "SI"],
    ["VINOS", "D.O. COMUNIDAD VALENCIANA", "PAGO DE LOS BALAGUESES, CHARDONNAY", "", "", "", "", 37, "", "SI"],
    
    ["VINOS", "D.O. RUEDA", "VERDEO, VERDEJO", "", "", "", 4.5, 19, "", "SI"],
    ["VINOS", "D.O. RUEDA", "VIÑA SALCEDA, VERDEJO", "", "", "", "", 22, "", "SI"],
    ["VINOS", "D.O. RUEDA", "TARSUS, VERDEJO", "", "", "", "", 25, "", "SI"],
    ["VINOS", "D.O. RUEDA", "MALCORTA, VERDEJO", "", "", "", "", 34, "", "SI"],
    ["VINOS", "D.O. RUEDA", "CAMPO ELISEO CUVEE ALEGRE, VERDEJO", "", "", "", "", 32, "", "SI"],

    ["VINOS", "D.O. RÍAS BAIXAS", "TERRAS GAUDAS, ALBARIÑO", "", "", "", "", 31, "", "SI"],
    ["VINOS", "D.O. RÍAS BAIXAS", "PAZO DAS BRUXAS, ALBARIÑO", "", "", "", "", 27, "", "SI"],
    ["VINOS", "D.O. RÍAS BAIXAS", "PAZO CILLEIRO, ALBARIÑO", "", "", "", "", 25, "", "SI"],
    ["VINOS", "D.O. RÍAS BAIXAS", "FILLABOA, ALBARIÑO", "", "", "", "", 34, "", "SI"],

    ["VINOS", "D.O.C. RIOJA", "AZPILICUETA, VIURA", "", "", "", "", 24, "", "SI"],
    ["VINOS", "D.O.C. RIOJA", "COLECCIÓN PRIVADA AZPILICUETA, VIURA", "", "", "", "", 36, "", "SI"],
    
    ["VINOS", "D.O. NAVARRA", "CHIVITE LAS FINCAS, CHARDONNAY, GARNACHA BLANCA", "", "", "", "", 21, "", "SI"],
    ["VINOS", "D.O. NAVARRA", "CHIVITE LEGARDETA, CHARDONNAY", "", "", "", "", 26, "", "SI"],
]

# Crear DataFrames
df_comida = pd.DataFrame(data_comida, columns=[
    'CATEGORIA', 'SUBCATEGORIA', 'NOMBRE_ES', 'NOMBRE_EN', 
    'DESC_ES', 'DESC_EN', 'PRECIO', 'ALERGENOS', 'VISIBLE'
])
df_bebida = pd.DataFrame(data_bebida, columns=[
    'CATEGORIA', 'SUBCATEGORIA', 'NOMBRE_ES', 'NOMBRE_EN', 
    'DESC_ES', 'DESC_EN', 'PRECIO_COPA', 'PRECIO_BOTELLA', 'ALERGENOS', 'VISIBLE'
])

# Guardar manteniendo configuración existente si es posible
# Como xlsxwriter reescribe, usaremos openpyxl para append o pandas ExcelWriter con mode='a' e if_sheet_exists='replace'
# Pero 'xlsxwriter' fue usado para crear formato rico. Pandas 'to_excel' puede borrar validaciones.
# Mejor estrategia: Usar openpyxl para insertar datos sin romper estilos.

try:
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    
    # 1. HUSA COMIDA
    if 'COMIDA' in wb.sheetnames:
        ws = wb['COMIDA']
        # Borrar datos viejos (filas > 1)
        # O simplemente hacer append. La hoja estaba vacía (solo headers).
        # Empezamos en fila 2
        for r_idx, row in enumerate(data_comida, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 2. HUSA BEBIDA
    if 'BEBIDAS' in wb.sheetnames:
        ws = wb['BEBIDAS']
        for r_idx, row in enumerate(data_bebida, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(file_path)
    print("Datos insertados correctamente en COMIDA y BEBIDAS.")

except Exception as e:
    print(f"Error insertando datos: {e}")
    # Fallback a pandas si openpyxl falla (aunque perderá formatos bonitos)
    print("Intentando fallback con pandas (puede perder estilos)...")
    with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
         df_comida.to_excel(writer, sheet_name='COMIDA', index=False, header=False, startrow=1)
         df_bebida.to_excel(writer, sheet_name='BEBIDAS', index=False, header=False, startrow=1)
