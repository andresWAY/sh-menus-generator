import pandas as pd
from openpyxl import load_workbook

file_path = "data/MENU_MASTER_FINAL.xlsx"

# Datos ficticios para Sandwiches (ya que el usuario no mandó foto de texto, pero pidió crearlos)
# Usamos una categoría nueva para que coja el fondo "SANDWICHES Y HAMBURGUESAS.jpg"
new_items = [
    # Cat, Subcat, NombreES, NombreEN, DescES, DescEN, Precio, Alergenos, Visible
    ["SANDWICHES Y HAMBURGUESAS", "", "CLUB SANDWICH", "CLUB SANDWICH", "POLLO, BACON, HUEVO, LECHUGA Y MAYONESA", "CHICKEN, BACON, EGG, LETTUCE & MAYO", 14, "1, 3, 7", "SI"],
    ["SANDWICHES Y HAMBURGUESAS", "", "HAMBURGUESA DE TERNERA", "BEEF BURGER", "QUESO CHEDDAR, CEBOLLA CARAMELIZADA Y PATATAS", "CHEDDAR CHEESE, CARAMELIZED ONION & FRIES", 16, "1, 7, 11", "SI"],
    ["SANDWICHES Y HAMBURGUESAS", "", "SANDWICH MIXTO", "HAM & CHEESE SANDWICH", "JAMÓN YORK Y QUESO", "BOILED HAM & CHEESE", 10, "1, 7", "SI"],
]

try:
    wb = load_workbook(file_path)
    if 'COMIDA' in wb.sheetnames:
        ws = wb['COMIDA']
        # Encontrar primera fila vacía
        first_empty_row = ws.max_row + 1
        
        for row_data in new_items:
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=first_empty_row, column=col_idx, value=value)
            first_empty_row += 1
            
    wb.save(file_path)
    print("Sandwiches añadidos al Excel.")
    
except Exception as e:
    print(f"Error añadiendo items: {e}")
